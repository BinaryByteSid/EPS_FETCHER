import io
import os
import time
from fastapi import FastAPI, HTTPException, Request, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import pandas as pd
from backend.scraper import (
    scrape_screener_quarters,
    filter_quarters,
    parse_quarter,
    resolve_company,
    get_bse_month_end_prices,
    generate_yoy_comparison,
    _isin_kind,
)

app = FastAPI(title="Quarterly EPS Fetcher & Exporter")

# Per-company wall-clock budget for pulling filed diluted EPS out of result
# PDFs. Generous by design: the user prefers correct filed figures over speed,
# and PDF downloads run in parallel, so a full 13-quarter range normally
# completes well within this. Quarters that still miss the budget fall back to
# Screener's basic EPS and are retried (and cached) on the next request.
PDF_EPS_BUDGET_SECONDS = 120.0

# The pdfplumber PDF fallback is the last resort and the main memory hog (it can
# OOM-kill a small instance). It is off by default so the deployed app stays
# crash-safe; set ENABLE_PDF_FALLBACK=1 to re-enable it locally. The BSE
# structured API and NSE XBRL cover the vast majority of companies without it.
ENABLE_PDF_FALLBACK = os.environ.get("ENABLE_PDF_FALLBACK", "").lower() in ("1", "true", "yes")

# Total wall-clock budget for a whole multi-company request. Without this cap a
# 10-company pivot would grant each company a fresh PDF budget and could block
# the single worker for many minutes, so Render's health checks fail and the
# instance is restarted (looks like a crash). Once this elapses, remaining
# companies fall back to Screener's figures (cheap) and fill in on a later
# request via the per-quarter caches.
MULTI_REQUEST_BUDGET_SECONDS = 200.0

# Enable CORS for development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Input Pydantic Model
class EPSRequest(BaseModel):
    symbols: list[str] = Field(None, description="List of NSE/BSE company symbols, e.g. ['RELIANCE', 'TCS']")
    symbol: str = Field(None, description="NSE/BSE company symbol fallback, e.g. RELIANCE")
    from_quarter: str = Field(..., description="Starting quarter, e.g. Mar 2023", example="Mar 2023")
    to_quarter: str = Field(..., description="Ending quarter, e.g. Mar 2026", example="Mar 2026")
    consolidated: bool = Field(True, description="Whether to fetch consolidated or standalone figures")
    eps_type: str = Field("diluted", description="EPS type to fetch: 'basic' or 'diluted'")
    period: str = Field("quarterly", description="Reporting period: 'quarterly' or 'yearly'")


# Cap on companies compared at once. Each company makes several rate-limited
# BSE calls, so a larger batch risks the request timing out.
MAX_COMPANIES = 10


def _normalize_symbol_inputs(req: EPSRequest) -> list[str]:
    symbol_list = req.symbols
    if not symbol_list and req.symbol:
        symbol_list = [req.symbol]
    if not symbol_list:
        raise HTTPException(status_code=400, detail="At least one company must be selected.")
    # De-duplicate (case-insensitively, preserving order) and cap at MAX_COMPANIES.
    seen = set()
    deduped = []
    for s in symbol_list:
        key = (s or "").strip().lower()
        if key and key not in seen:
            seen.add(key)
            deduped.append(s.strip())
    return deduped[:MAX_COMPANIES]


def _free_float_value(resolved: dict, deadline: float = None):
    """Latest public-shareholding % for a company, or None. Cheap and cached."""
    try:
        from backend.bse_results_api import fetch_free_float
        ff = fetch_free_float(resolved.get("bse", ""), deadline=deadline)
        return ff["free_float"] if ff else None
    except Exception as e:
        print(f"Warning: free-float fetch failed for {resolved.get('symbol')}: {e}")
        return None


def _build_row_from_query(query: str, filtered: list[dict], field: str = "EPS in Rs") -> dict:
    """
    Builds a pivoted row (company metadata + one value per quarter) for the given
    record field — "EPS in Rs" for the EPS pivot, "Net Profit" for the PAT pivot.
    """
    meta = resolve_company(query)
    query_clean = query.strip().upper()
    isin_value = meta["isin"] or (query_clean if len(query_clean) == 12 and query_clean.startswith("IN") else "N/A")
    row_dict = {
        "Symbol": meta["symbol"],
        "Name": meta.get("name") or meta["symbol"],
        "BSE": meta["bse"] or "N/A",
        "ISIN": isin_value
    }

    for rec in filtered:
        q = rec["Quarter"]
        row_dict[q] = rec.get(field)

    return row_dict


def scrape_and_enrich_quarters(resolved: dict, consolidated: bool, from_quarter: str, to_quarter: str, deadline: float = None, eps_type: str = "diluted", period: str = "quarterly") -> tuple[list[dict], list[dict]]:
    """
    Scrapes Screener (quarterly or yearly per `period`), fetches basic or diluted
    EPS for the requested range and their YoY previous periods, enriches them,
    and returns (all_records, filtered_records).
    """
    from backend.bse_eps import fetch_diluted_eps_bse
    from backend.nse_eps import fetch_diluted_eps_nse
    from backend.bse_results_api import fetch_diluted_eps_bse_api

    eps_type = (eps_type or "diluted").lower()
    yearly = (period or "quarterly").lower() in ("yearly", "annual", "year")

    # A non-equity ISIN survived resolution unchanged (mutual fund / other) — it
    # is not a listed company, so fail fast with a clear, actionable message
    # instead of a generic "not found on Screener".
    kind = _isin_kind(resolved.get("symbol", ""))
    if kind == "fund":
        raise Exception(f"'{resolved['symbol']}' is a mutual-fund ISIN (INF…). This tool covers listed companies only; equity ISINs start with 'INE'.")
    if kind == "other":
        raise Exception(f"'{resolved['symbol']}' is a non-equity ISIN. This tool covers listed companies only (equity ISINs start with 'INE').")

    records = scrape_screener_quarters(resolved["symbol"], consolidated, period=period)
    filtered = filter_quarters(records, from_quarter, to_quarter)

    # Identify quarters to enrich: filtered quarters + YoY previous quarters
    quarters_to_enrich = set()
    for r in filtered:
        quarters_to_enrich.add(r["Quarter"])
        parts = r["Quarter"].strip().split()
        if len(parts) == 2:
            month, year_str = parts
            try:
                prev_q = f"{month} {int(year_str) - 1}"
                quarters_to_enrich.add(prev_q)
            except ValueError:
                pass

    # Parse quarters to find min and max boundaries for fetch_diluted_eps_bse
    parsed_to_enrich = []
    for q in quarters_to_enrich:
        try:
            parsed_to_enrich.append(parse_quarter(q))
        except ValueError:
            pass

    if parsed_to_enrich:
        min_q = min(parsed_to_enrich)
        max_q = max(parsed_to_enrich)

        eps_map = {}

        def _still_missing() -> bool:
            return any(q not in eps_map for q in quarters_to_enrich)

        # Primary source: BSE's structured results API (quarterly or annual).
        try:
            api_map = fetch_diluted_eps_bse_api(
                resolved.get("bse", ""), resolved.get("name", resolved["symbol"]),
                min_q, max_q, consolidated=consolidated, deadline=deadline,
                eps_type=eps_type, period=period
            )
            eps_map.update(api_map)
        except Exception as e:
            print(f"Warning: BSE results-API EPS fetch failed for {resolved['symbol']}: {e}")

        # The NSE-XBRL and BSE-PDF fallbacks parse per-quarter filings, so they
        # only apply to quarterly mode. For yearly, unresolved years fall back to
        # Screener's annual figures already present on the record.
        if not yearly and _still_missing():
            try:
                nse_map = fetch_diluted_eps_nse(
                    resolved["symbol"], min_q, max_q,
                    consolidated=consolidated, deadline=deadline,
                    eps_type=eps_type
                )
                for q_label, info in nse_map.items():
                    eps_map.setdefault(q_label, info)
            except Exception as e:
                print(f"Warning: NSE XBRL EPS fetch failed for {resolved['symbol']}: {e}")

        if ENABLE_PDF_FALLBACK and not yearly and _still_missing():
            expected_basic = {
                r["Quarter"]: r["EPS in Rs"] for r in records
                if isinstance(r.get("EPS in Rs"), (int, float))
            }
            try:
                bse_map = fetch_diluted_eps_bse(
                    company=resolved,
                    from_q=min_q,
                    to_q=max_q,
                    consolidated=consolidated,
                    deadline=deadline,
                    expected_basic=expected_basic,
                    eps_type=eps_type
                )
                for q_label, info in bse_map.items():
                    eps_map.setdefault(q_label, info)
            except Exception as e:
                print(f"Warning: Failed to fetch PDF EPS for {resolved['symbol']}: {e}")

        # Enrich all records in records list
        for r in records:
            q = r["Quarter"]
            info = eps_map.get(q)
            if info is not None and info.get("value") is not None:
                r["EPS in Rs"] = info["value"]
                r["EPS Source"] = info.get("source") or ("NSE XBRL" if info.get("xbrl") else "BSE PDF")
            else:
                r["EPS Source"] = "screener (basic)" if eps_type == "basic" else "screener"

            # Filed Profit After Tax rides along in the same BSE-API filing as
            # the diluted EPS; override Screener's net profit with it when present.
            if info is not None and info.get("pat") is not None:
                r["Net Profit"] = info["pat"]
                r["PAT Source"] = info.get("source") or "BSE API"
            else:
                r["PAT Source"] = "screener"
    else:
        for r in records:
            r["EPS Source"] = "screener (basic)" if eps_type == "basic" else "screener"
            r["PAT Source"] = "screener"

    # Re-filter to get updated filtered records
    filtered_enriched = filter_quarters(records, from_quarter, to_quarter)

    return records, filtered_enriched


def _build_metadata_row(query: str, meta: dict) -> dict:
    query_clean = query.strip().upper()
    return {
        "Symbol": meta["symbol"],
        "BSE": meta["bse"] or "N/A",
        "ISIN": meta["isin"] or (query_clean if len(query_clean) == 12 and query_clean.startswith("IN") else "N/A")
    }


def _style_excel_sheet(worksheet, df: pd.DataFrame) -> None:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=11, bold=True)
    regular_font = Font(name="Segoe UI", size=11)

    thin = Side(border_style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    num_cols = len(df.columns)

    for col in range(1, num_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 28

    for row in range(2, worksheet.max_row + 1):
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = regular_font
            cell.border = border

            if col in [1, 2, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col == 1:
                    cell.font = bold_font
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    for col in worksheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 13)

@app.get("/api/companies")
async def get_companies():
    """
    Returns the loaded database of companies from ISIN'S.xlsx for autocomplete search.
    """
    from backend.scraper import COMPANIES_DB
    return COMPANIES_DB

@app.post("/api/fetch-eps")
def fetch_eps(req: EPSRequest):
    """
    Fetches quarterly financials for multiple companies, filters them,
    pivots them (quarters in horizontal, companies in vertical) and returns JSON.
    """
    try:
        symbol_list = _normalize_symbol_inputs(req)

        rows = []
        pat_rows = []
        company_entries = []
        all_quarters_set = set()
        warnings = []

        # Shared budget across all companies bounds the whole request; once spent,
        # later companies get an expired deadline (Screener-only, fast) so the
        # worker is never blocked long enough to trip Render's health checks.
        overall_deadline = time.monotonic() + MULTI_REQUEST_BUDGET_SECONDS

        for query in symbol_list:
            try:
                deadline = min(time.monotonic() + PDF_EPS_BUDGET_SECONDS, overall_deadline)
                resolved = resolve_company(query)
                # Free float first (before EPS enrichment throttles BSE).
                ff_val = _free_float_value(resolved, deadline=min(time.monotonic() + 12, overall_deadline))
                records, filtered = scrape_and_enrich_quarters(resolved, req.consolidated, req.from_quarter, req.to_quarter, deadline=deadline, eps_type=req.eps_type, period=req.period)

                if not filtered:
                    continue

                row_dict = _build_row_from_query(query, filtered)
                pat_row = _build_row_from_query(query, filtered, field="Net Profit")
                row_dict["Free Float %"] = ff_val
                pat_row["Free Float %"] = ff_val
                company_entries.append({
                    "query": query,
                    "resolved": resolved,
                    "filtered": filtered,
                    "free_float": ff_val,
                })
                for rec in filtered:
                    all_quarters_set.add(rec["Quarter"])
                rows.append(row_dict)
                pat_rows.append(pat_row)
            except Exception as e:
                warnings.append(f"Error fetching '{query}': {str(e)}")

        if not rows:
            raise Exception("Failed to fetch data for all requested companies:\n" + "\n".join(warnings))

        # Sort quarters chronologically
        sorted_quarters = sorted(list(all_quarters_set), key=parse_quarter)

        return {
            "status": "success",
            "quarters": sorted_quarters,
            "rows": rows,
            "pat_rows": pat_rows,
            "warnings": warnings
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/export-excel")
def export_excel(req: EPSRequest):
    """
    Fetches quarterly financials for multiple companies, pivots them,
    and exports to a single pivoted Excel worksheet.
    """
    try:
        symbol_list = _normalize_symbol_inputs(req)

        rows = []
        pat_rows = []
        company_entries = []
        all_quarters_set = set()
        warnings = []

        # Shared budget across all companies bounds the whole request; once spent,
        # later companies get an expired deadline (Screener-only, fast) so the
        # worker is never blocked long enough to trip Render's health checks.
        overall_deadline = time.monotonic() + MULTI_REQUEST_BUDGET_SECONDS

        for query in symbol_list:
            try:
                deadline = min(time.monotonic() + PDF_EPS_BUDGET_SECONDS, overall_deadline)
                resolved = resolve_company(query)
                # Free float first (before EPS enrichment throttles BSE).
                ff_val = _free_float_value(resolved, deadline=min(time.monotonic() + 12, overall_deadline))
                records, filtered = scrape_and_enrich_quarters(resolved, req.consolidated, req.from_quarter, req.to_quarter, deadline=deadline, eps_type=req.eps_type, period=req.period)

                if not filtered:
                    continue

                row_dict = _build_row_from_query(query, filtered)
                pat_row = _build_row_from_query(query, filtered, field="Net Profit")
                row_dict["Free Float %"] = ff_val
                pat_row["Free Float %"] = ff_val
                company_entries.append({
                    "query": query,
                    "resolved": resolved,
                    "filtered": filtered,
                    "free_float": ff_val,
                })
                for rec in filtered:
                    all_quarters_set.add(rec["Quarter"])
                rows.append(row_dict)
                pat_rows.append(pat_row)
            except Exception as e:
                warnings.append(f"Error fetching '{query}': {str(e)}")

        if not rows:
            raise Exception("No data available to export:\n" + "\n".join(warnings))

        # Sort quarters chronologically
        sorted_quarters = sorted(list(all_quarters_set), key=parse_quarter)

        stock_rows = []
        for entry in company_entries:
            stock_row = _build_metadata_row(entry["query"], entry["resolved"])
            stock_row["Free Float %"] = entry.get("free_float")
            price_map = get_bse_month_end_prices(entry["resolved"], sorted_quarters)
            stock_row.update(price_map)
            stock_rows.append(stock_row)

        # Create DataFrames
        df = pd.DataFrame(rows)
        pat_df = pd.DataFrame(pat_rows)
        price_df = pd.DataFrame(stock_rows)

        # Enforce column order: Symbol, BSE, ISIN, Free Float %, then quarters
        cols_order = ["Symbol", "BSE", "ISIN", "Free Float %"] + sorted_quarters

        def _shape(frame):
            for col in cols_order:
                if col not in frame.columns:
                    frame[col] = None
            frame = frame[cols_order]
            frame.rename(columns={"BSE": "BSE Code", "ISIN": "ISIN"}, inplace=True)
            return frame

        df = _shape(df)
        pat_df = _shape(pat_df)
        price_df = _shape(price_df)

        period_word = "Yearly" if (req.period or "").lower() in ("yearly", "annual", "year") else "Quarterly"
        eps_sheet = f"{period_word} EPS ({'Basic' if req.eps_type == 'basic' else 'Diluted'})"
        pat_sheet = f"{period_word} PAT (Cr)"
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=eps_sheet)
            pat_df.to_excel(writer, index=False, sheet_name=pat_sheet)
            price_df.to_excel(writer, index=False, sheet_name='Stock Prices')

            _style_excel_sheet(writer.sheets[eps_sheet], df)
            _style_excel_sheet(writer.sheets[pat_sheet], pat_df)
            _style_excel_sheet(writer.sheets['Stock Prices'], price_df)

        output.seek(0)

        eps_label = req.eps_type if req.eps_type in ('basic', 'diluted') else 'diluted'
        filename = f"{period_word}_EPS_{eps_label}_{'consolidated' if req.consolidated else 'standalone'}_{req.from_quarter.replace(' ', '')}_to_{req.to_quarter.replace(' ', '')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/resolve")
def api_resolve(query: str):
    """
    Resolves a user-entered symbol, BSE code, or ISIN to a company metadata record.
    """
    try:
        resolved = resolve_company(query)
        return resolved
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/scrape")
def api_scrape(
    symbol: str,
    consolidated: bool = True,
    from_quarter: str = None,
    to_quarter: str = None,
    eps_type: str = "diluted",
    period: str = "quarterly"
):
    """
    Scrapes quarterly or yearly financials for a single company, generates
    comparisons, and fetches BSE stock prices.
    """
    try:
        resolved = resolve_company(symbol)

        # Free float (public shareholding %) — fetched up front, before the heavy
        # EPS enrichment hammers (and gets throttled by) the same BSE host.
        free_float = None
        try:
            from backend.bse_results_api import fetch_free_float
            free_float = fetch_free_float(resolved.get("bse", ""), deadline=time.monotonic() + 10)
        except Exception as e:
            print(f"Warning: free-float fetch failed for {resolved['symbol']}: {e}")

        raw_records = scrape_screener_quarters(resolved["symbol"], consolidated, period=period)

        # Determine available periods
        available_quarters = [r["Quarter"] for r in raw_records]

        # If from/to are not specified, default to full range
        if not from_quarter and available_quarters:
            from_quarter = available_quarters[0]
        if not to_quarter and available_quarters:
            to_quarter = available_quarters[-1]

        deadline = time.monotonic() + PDF_EPS_BUDGET_SECONDS

        records, filtered = scrape_and_enrich_quarters(resolved, consolidated, from_quarter, to_quarter, deadline=deadline, eps_type=eps_type, period=period)
        comparison = generate_yoy_comparison(records, filtered)
        
        # Fetch month-end stock prices
        quarters_to_fetch = [r["Quarter"] for r in filtered]
        prices = get_bse_month_end_prices(resolved, quarters_to_fetch)
        
        # Format warnings/notes
        notes = []
        for q in quarters_to_fetch:
            if prices.get(q) is None:
                notes.append(f"Month-end stock price for {q} not found on BSE.")
                
        return {
            "company": resolved,
            "records": filtered,
            "comparison": comparison,
            "prices": prices,
            "notes": notes,
            "eps_type": eps_type,
            "period": period,
            "free_float": free_float
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/pivot")
def api_pivot(req: EPSRequest):
    """
    Alias / implementation for POST /api/pivot requested by the frontend.
    """
    return fetch_eps(req)

@app.get("/api/export")
def api_export(
    symbols: list[str] = Query(...),
    from_quarter: str = Query(...),
    to_quarter: str = Query(...),
    consolidated: bool = Query(True),
    eps_type: str = Query("diluted"),
    period: str = Query("quarterly")
):
    """
    FastAPI GET endpoint for exporting pivoted Excel data.
    """
    # Create an EPSRequest to reuse the export_excel logic
    req = EPSRequest(
        symbols=symbols,
        from_quarter=from_quarter,
        to_quarter=to_quarter,
        consolidated=consolidated,
        eps_type=eps_type,
        period=period
    )
    return export_excel(req)

# Mount frontend directory for static assets (/static/app.js, style.css)
# Use safe pathing that works across OSes
frontend_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "frontend")
if os.path.exists(frontend_dir):
    app.mount("/static", StaticFiles(directory=frontend_dir), name="static")

# Serve index.html at root URL
@app.get("/", response_class=HTMLResponse)
async def get_index():
    index_path = os.path.join(frontend_dir, "index.html")
    if not os.path.exists(index_path):
        return HTMLResponse(content="<h1>Frontend index.html not found! Check directory structure.</h1>", status_code=404)
    
    with open(index_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())
